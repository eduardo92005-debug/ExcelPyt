import openpyxl as pyx
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from pathlib import Path
import pyexcel

SEPARATOR = ':'
START_ROW_ELET = "22"
START_ROW_DISC = "27"


def getColumnValues(datasheet, type_row, time_values, first_term, ratio):
    col_values = []
    number_test = datasheet["B3"]
    if(type_row == "elet"):
        col_number = getColumnByNumber(first_term, ratio, time_values)
        startsrc_index = col_number+START_ROW_ELET
        finalsrc_index = col_number+str((int(START_ROW_ELET)+2))
        search_index = startsrc_index+SEPARATOR+finalsrc_index
        for index in range(0, 3):
            for col in datasheet[search_index][index]:
                if(col.value != None):
                    col_values.append(col.value)
    else:
        col_number = getColumnByNumber(first_term, ratio, time_values)
        startsrc_index = col_number+START_ROW_DISC
        finalsrc_index = col_number+str((int(START_ROW_DISC)+2))
        search_index = startsrc_index+SEPARATOR+finalsrc_index
        for index in range(0, 2):
            for col in datasheet[search_index][index]:
                if(col.value != None):
                    col_values.append(col.value)
    return [number_test, col_values]


def saveExcelFile(csv_name='./outputs/output1261.csv'):
    sheet = pyexcel.get_sheet(file_name=csv_name, delimiter=";")
    sheet.save_as(".mytemp.xlsx")


def loadBook(name='.mytemp.xlsx'):
    book = pyx.load_workbook(name)
    return book


def getColumnByNumber(first_term, ratio, value_to_encounter_column):
    n_search_term = round(
        (value_to_encounter_column - first_term + ratio)/ratio)
    column_index = get_column_letter(n_search_term+1)
    return column_index


def getArcValues(type_row, time_values=30, first_term=15, ratio=5):
    path = Path('outputs/')
    test_files = list(path.glob('output*.csv'))
    count_files = len(test_files)
    arc_values = []
    for index in range(0, count_files):
        saveExcelFile(r'./'+str(test_files[index]))
        book = loadBook()
        ws = book.active
        arc_values.append(getColumnValues(ws,type_row,time_values,first_term,ratio))
    return arc_values



